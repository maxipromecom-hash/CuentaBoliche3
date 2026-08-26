def exportar(rows, destino, nombre, sector_nombre="Todos los sectores", sector_rows=None):
    # Import diferido: openpyxl solo se carga cuando el usuario exporta.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Control"

    headers = [
        "Producto", "Categoría", "Inicial", "Ingresos", "Transf. ingreso",
        "Salidas", "Mermas", "Transf. salida", "Esperado", "Físico",
        "Diferencia", "Faltante", "Costo", "Pérdida", "Estado"
    ]
    ws.append(["CONTROL DE STOCK Y PÉRDIDAS"])
    ws.append([nombre])
    ws.append([f"Sector: {sector_nombre}"])
    ws.append([])
    ws.append(headers)

    dark = "1F2937"
    red = "FEE2E2"
    yellow = "FEF3C7"
    green = "DCFCE7"
    thin = Side(style="thin", color="D1D5DB")

    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].alignment = Alignment(horizontal="center")
    for rr in [2, 3]:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=len(headers))
        ws.cell(rr, 1).font = Font(bold=True, size=12)
        ws.cell(rr, 1).alignment = Alignment(horizontal="center")

    for c in ws[5]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=thin)

    for r in rows:
        if r["final"] is None:
            estado = "PENDIENTE"
        elif (r["faltante"] or 0) > 0:
            estado = "DIFERENCIA"
        elif (r["diferencia"] or 0) > 0:
            estado = "SOBRANTE"
        else:
            estado = "CONTROLADO"
        ws.append([
            r["producto"], r["categoria"], r["inicial"], r["ingresos"], r["transferencias_in"],
            r["salidas"], r["mermas"], r["transferencias_out"], r["esperado"],
            "SIN CONTAR" if r["final"] is None else r["final"],
            "" if r["diferencia"] is None else r["diferencia"],
            "" if r["faltante"] is None else r["faltante"], r["costo"],
            "" if r["perdida"] is None else r["perdida"], estado
        ])
        fill = green if estado == "CONTROLADO" else yellow if estado in ("PENDIENTE", "SOBRANTE") else red
        ws.cell(ws.max_row, 15).fill = PatternFill("solid", fgColor=fill)

    ws.append([])
    ws.append(["PÉRDIDA TOTAL", sum((r["perdida"] or 0) for r in rows)])
    ws.append(["PENDIENTES", sum(1 for r in rows if r["final"] is None)])
    ws.append(["PRODUCTOS CON DIFERENCIA", sum(1 for r in rows if (r["faltante"] or 0) > 0)])

    for row in ws.iter_rows(min_row=6):
        for cell in row:
            cell.border = Border(bottom=thin)
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 11), 35)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:O{max(5, 5 + len(rows))}"

    # Hoja adicional con el estado de cada sector.
    if sector_rows:
        ss = wb.create_sheet("Sectores")
        ss.append(["CONTROL POR SECTOR"])
        ss.merge_cells("A1:F1")
        ss["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ss["A1"].fill = PatternFill("solid", fgColor=dark)
        ss["A1"].alignment = Alignment(horizontal="center")
        ss.append(["Sector", "Productos", "Contados", "Pendientes", "Con diferencia", "Pérdida"])
        for c in ss[2]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=dark)
        for s in sector_rows:
            ss.append([s["sector"], s["productos"], s["contados"], s["pendientes"], s["diferencias"], s["perdida"]])
        for col in range(1, 7):
            ss.column_dimensions[get_column_letter(col)].width = 20
        ss.freeze_panes = "A3"

    wb.save(destino)
